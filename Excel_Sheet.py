import openpyxl
file = " file path"
workbook=openpyxl.load_workbook(file)

sh1=workbook["sheet1"]

rows = sh1.max_row
colums = sh1.max_column

# Reading excel data
for r in rows:
    for c in colums:
        print(sh1.cell(r,c).value,end='')
    print()

# Writting into excel data
for r in rows:
    for c in colums:
        sh1.cell(r,c).value="end"
workbook.save(file)