from selenium import webdriver
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By

# Chrome driver setup automate
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

# firefox driver
# from selenium.webdriver.firefox.service import Service as FirefoxService
# from webdriver_manager.firefox import GeckoDriverManager
# driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()))

# Edge setup
# from selenium.webdriver.edge.service import Service as EdgeService
# from webdriver_manager.microsoft import EdgeChromiumDriverManager
# driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))


driver.get("https://www.yatra.com/")
driver.maximize_window()
# driver.implicitly_wait(10)
wait = WebDriverWait(driver, 30)
wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@if='BE_flight_origin_date'] "))).click()
all_dates= wait.until(expected_conditions.element_to_be_clickable((By.XPATH,"//div[@id='monthWrapper']//tbody//td[@class!='inActiveTD']"))).find_elements(By.XPATH,"//div[@id='monthWrapper']//tbody//td[@class!='inActiveTD']")

for date in all_dates:
    if date.get_attribute("date-date") == "24/12/2022":
        date.click()
        break
























# # Read xml file
# file = "E:\data.xlsx"
# # workbook open
# workbook = openpyxl.load_workbook(file)
# # sheet access
# sh1 = workbook['Sheet1']
# # row
# rows = sh1.max_row
# # column
# col = sh1.max_column
# 
# # reading all rows and colunmns
# for r in range(1, rows):
#     for c in range(1, col):
#         print(sh1.cell(r, c).value, end="")
#     print(" ")
# 
# # writting into excel sheet
# for r in range(8, 10):
#     for c in range(10, 16):
#         sh1.cell(r, c).value = "welcome"
# 
# workbook.save(file)
